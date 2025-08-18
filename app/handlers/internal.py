from fastapi import HTTPException, UploadFile
from sqlalchemy.orm import Session
from app import models, schemas, crud
from app.handlers.external import fetch_books_from_google
import pandas as pd
import os
import tempfile
from fastapi.responses import FileResponse
from openpyxl import Workbook, load_workbook
from app.decorator import measure_performance
from io import BytesIO

#на пандас тут все
@measure_performance
async def export_books_handler(db: Session, file_format: str = "xlsx"):
    books = db.query(models.Book).all()
    if not books:
        raise ValueError("Нет данных для экспорта")

    data = [
        {
            "ID": b.id,
            "title": b.title,
            "author": b.author,
            "year": b.year,
        }
        for b in books
    ]

    df = pd.DataFrame(data)

    tmp_dir = tempfile.gettempdir()
    file_path = os.path.join(tmp_dir, f"books_export.{file_format}")

    df.to_excel(file_path, index=False, engine="openpyxl")

    return FileResponse(
        file_path,
        filename=f"books_export.{file_format}",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@measure_performance
async def import_books_from_excel(file: UploadFile, db: Session):
    contents = await file.read()
    df = pd.read_excel(BytesIO(contents))

    required_columns = {"title", "author", "year"}
    if not required_columns.issubset(df.columns):
        missing = required_columns - set(df.columns)
        raise ValueError(f"Отсутствуют колонки: {', '.join(missing)}")

    imported_count = 0

    for _, row in df.iterrows():
        book = models.Book(
            title=row["title"],
            author=row["author"],
            year=int(row["year"]) if not pd.isna(row["year"]) else None
        )
        db.add(book)
        imported_count += 1

    db.commit()
    return imported_count

#тут на openpyxl все
@measure_performance
async def export_books_handler_openpyxl(db: Session, file_format: str = "xlsx"):
    books = db.query(models.Book).all()
    if not books:
        raise ValueError("Нет данных для экспорта")

    wb = Workbook()
    ws = wb.active
    ws.append(["ID", "title", "author", "year"])

    for b in books:
        ws.append([b.id, b.title, b.author, b.year])

    tmp_dir = tempfile.gettempdir()
    file_path = os.path.join(tmp_dir, f"books_export.{file_format}")
    wb.save(file_path)

    return FileResponse(
        file_path,
        filename=f"books_export.{file_format}",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@measure_performance
async def import_books_from_openpyxl(file: UploadFile, db: Session):
    contents = await file.read()
    wb = load_workbook(filename=BytesIO(contents), data_only=True)
    ws = wb.active

    headers = [cell.value.lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    required_columns = {"title", "author", "year"}
    if not required_columns.issubset(headers):
        missing = required_columns - set(headers)
        raise ValueError(f"Отсутствуют колонки: {', '.join(missing)}")

    imported_count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        book = models.Book(
            title=row_data["title"],
            author=row_data["author"],
            year=int(row_data["year"]) if row_data["year"] else None
        )
        db.add(book)
        imported_count += 1

    db.commit()
    return imported_count

async def get_books_by_filters(db: Session, filters, skip: int, limit: int):
    query = db.query(models.Book)

    filter_mapping = {
        "book_id": lambda v: models.Book.id == v,
        "title": lambda v: models.Book.title.ilike(f"%{v}%"),
        "author": lambda v: models.Book.author.ilike(f"%{v}%"),
        "year": lambda v: models.Book.year == v,
    }

    for field, condition in filter_mapping.items():
        value = getattr(filters, field, None)
        if value is not None:
            query = query.filter(condition(value))


    results = query.offset(skip).limit(limit).all()

    if results:
        return results

    if filters.title or filters.author or filters.year:
        return await fetch_books_from_google(
            title=filters.title,
            author=filters.author,
            year=filters.year,
            limit=limit
        )

    raise HTTPException(status_code=404, detail="Книги не найдены")


def create_book_handler(book: schemas.BookCreate, db: Session) -> schemas.BookOut:
    if not book.title or not book.author:
        raise HTTPException(status_code=422, detail="Название книги и автор обязательны")

    new_book = models.Book(
        title=book.title,
        author=book.author,
        year=book.year
    )
    db.add(new_book)
    db.commit()
    db.refresh(new_book)
    return new_book

def update_book_handler(book_id: str, book: schemas.BookCreate, db: Session):
    updated = crud.update_book(db, book_id, book)
    if updated is None:
        raise HTTPException(status_code=404, detail="Книга не найдена")
    return updated


def delete_book_handler(book_id: str, db: Session):
    deleted = crud.delete_book(db, book_id)
    if deleted is None:
        raise HTTPException(status_code=404, detail="Книга не найдена")
    return {"message": "Книга удалена"}
